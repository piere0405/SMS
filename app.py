# -*- coding: utf-8 -*-
"""
GENERADOR INTELIGENTE DE CAMPAÑAS DE ACTIVACIÓN — BBVA / PlusMetas
==================================================================
Flujo: Subir FORMALIZADAS + ACTIVADAS -> cruzar por DNI -> No Activadas
-> (segmento tarjeta) -> primeros 999 -> analizar promociones
-> alertas + recomendación por promoción -> speech <=160 sin tildes
-> Excel (hoja REC OTROS, formato plantilla).

Ejecutar:  streamlit run app.py
"""

import io
import re
import unicodedata
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# CONFIG GLOBAL / BRANDING PLUSMETAS
# ============================================================================
NAVY = "#145078"; LIME = "#AFCB07"
LIMITE_CARACTERES = 160
CIERRE_OBLIGATORIO = "activa hoy por la app BBVA o cajero mas cercano"
HOJA_SALIDA = "REC OTROS"
LIMITE_REGISTROS = 999
DEFAULT_HOJA_ACTIVADAS = "DTA JULIO"

# Anchos exactos de la plantilla sms (REC OTROS)
ANCHOS = {"A": 12.42578125, "B": 124.28515625, "C": 11.5703125}

st.set_page_config(page_title="Generador de Campañas de Activación · PlusMetas / BBVA",
                   page_icon="📲", layout="wide")

st.markdown(f"""
<style>
.block-container {{ padding-top:1.2rem; max-width:1250px; }}
.pm-header {{ background:linear-gradient(100deg,{NAVY} 0%,#0d3a5a 60%,{NAVY} 100%);
  border-radius:14px; padding:22px 26px; margin-bottom:18px; display:flex;
  align-items:center; justify-content:space-between; box-shadow:0 6px 20px rgba(20,80,120,.25); }}
.pm-logo {{ font-weight:800; font-size:26px; color:#fff; }} .pm-logo span {{ color:{LIME}; }}
.pm-header .sub {{ color:#dbe9f2; font-size:13px; margin-top:2px; }}
.pm-badge {{ background:{LIME}; color:{NAVY}; font-weight:800; font-size:12px; padding:6px 14px; border-radius:20px; }}
.kpi {{ background:#fff; border:1px solid #e6eef4; border-left:5px solid {NAVY}; border-radius:12px;
  padding:14px 16px; box-shadow:0 2px 8px rgba(0,0,0,.04); }}
.kpi .lbl {{ font-size:12px; color:#6b7b88; text-transform:uppercase; letter-spacing:.5px; }}
.kpi .val {{ font-size:28px; font-weight:800; color:{NAVY}; line-height:1.1; }}
.kpi.lime {{ border-left-color:{LIME}; }} .kpi.lime .val {{ color:#6f8100; }}
.pill {{ display:inline-block; padding:3px 12px; border-radius:20px; font-size:12px; font-weight:700; }}
.pill.red{{background:#fdecec;color:#c0261a;}} .pill.orange{{background:#fff3e2;color:#b96a00;}}
.pill.green{{background:#e9f6ec;color:#1f8a3b;}} .pill.yellow{{background:#fef9e0;color:#8a7400;}}
.pill.blue{{background:#e7f0fb;color:#1358a8;}}
.promo {{ background:#fff; border:1px solid #e2ecf3; border-left:5px solid {NAVY}; border-radius:12px;
  padding:14px 18px; margin:10px 0; box-shadow:0 2px 8px rgba(0,0,0,.04); }}
.promo.psi {{ border-left-color:#b96a00; }}
.promo h4 {{ margin:0 0 4px; color:{NAVY}; }}
.sectionbar {{ background:{NAVY}; color:#fff; font-weight:700; padding:9px 16px; border-radius:8px;
  margin:18px 0 10px; font-size:15px; }}
div.stButton>button, div.stDownloadButton>button {{ background:{NAVY}; color:#fff; border:none;
  border-radius:8px; font-weight:700; padding:10px 22px; }}
div.stButton>button:hover, div.stDownloadButton>button:hover {{ background:#0d3a5a; color:{LIME}; }}
.small {{ font-size:12px; color:#6b7b88; }}
</style>""", unsafe_allow_html=True)


# ============================================================================
# UTILIDADES
# ============================================================================
def quitar_tildes(texto: str) -> str:
    if texto is None: return ""
    texto = str(texto).replace("ñ", "\uE000").replace("Ñ", "\uE001")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.replace("\uE000", "ñ").replace("\uE001", "Ñ")

def normaliza_dni(serie: pd.Series) -> pd.Series:
    return (serie.astype(str).str.strip()
            .str.replace(r"\.0$", "", regex=True).str.replace(r"\D", "", regex=True))

def normaliza_celular(valor, prefijo="51") -> str:
    d = re.sub(r"\D", "", str(valor))
    if not d: return ""
    d = d.lstrip("0")
    if d.startswith(prefijo) and len(d) == len(prefijo) + 9: return d
    return prefijo + d

PARTICULAS = {"DE","DEL","LA","LAS","LOS","SAN","SANTA","MC","MAC","VON","DA","DI"}

def primer_nombre(nombre_completo: str) -> str:
    """Base BBVA: APELLIDO APELLIDO NOMBRE... -> primer nombre = token tras 2 apellidos."""
    if not nombre_completo or pd.isna(nombre_completo): return ""
    toks = [t for t in re.split(r"\s+", str(nombre_completo).strip()) if t]
    if not toks: return ""
    i, apellidos = 0, 0
    while i < len(toks) and apellidos < 2:
        if toks[i].upper() in PARTICULAS and i + 1 < len(toks):
            i += 2
        else:
            i += 1
        apellidos += 1
    return (toks[i] if i < len(toks) else toks[-1]).upper()


# ============================================================================
# LECTURA EXCEL / DETECCIÓN DE COLUMNAS
# ============================================================================
ALIAS = {
    "DNI":     ["DNI","DOC","DOCUMENTO","NRODOC","NRO_DOC","NUM_DOC","NUMERO_DOCUMENTO"],
    "CLIENTE": ["CLIENTE","NOMBRE","NOMBRES","NOMBRE_COMPLETO","NOMBRECOMPLETO",
                "NOMBRE COMPLETO","APELLIDOS Y NOMBRES","NOMBRE_CLIENTE"],
    "CELULAR": ["CELULAR","CELL","TELEFONO","CEL","MOVIL","NUMERO","TELF","TELEFONO1"],
    "TARJETA": ["TARJETA_WF","TARJETA WF"],
    "FECHA":   ["FECHA_FORMALIZACION","FECHA FORMALIZACION","FEC_FORMALIZACION","FECHA_FORMALIZ"],
}

def _match_col(cols, alias_list):
    up = {str(c).strip().upper(): c for c in cols}
    for a in alias_list:
        if a in up: return up[a]
    for a in alias_list:
        for k, orig in up.items():
            if a in k: return orig
    return None

@st.cache_data(show_spinner=False)
def leer_libro(file_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    return {sh: xls.parse(sh) for sh in xls.sheet_names}

def hojas_validas_formalizadas(libro):
    return [sh for sh, df in libro.items()
            if df is not None and not df.empty
            and _match_col(df.columns, ALIAS["DNI"])
            and _match_col(df.columns, ALIAS["CLIENTE"])
            and _match_col(df.columns, ALIAS["CELULAR"])]

def hojas_con_dni(libro):
    return [sh for sh, df in libro.items()
            if df is not None and not df.empty and _match_col(df.columns, ALIAS["DNI"])]


# ============================================================================
# CRUCE
# ============================================================================
def construir_formalizadas(libro, hojas):
    frames, avisos = [], []
    for sh in hojas:
        df = libro[sh]
        cD = _match_col(df.columns, ALIAS["DNI"])
        cC = _match_col(df.columns, ALIAS["CLIENTE"])
        cT = _match_col(df.columns, ALIAS["CELULAR"])
        cK = _match_col(df.columns, ALIAS["TARJETA"])
        cF = _match_col(df.columns, ALIAS["FECHA"])
        if not (cD and cC and cT):
            faltan = [k for k, c in [("DNI", cD), ("CLIENTE", cC), ("CELULAR", cT)] if not c]
            avisos.append(f"Hoja '{sh}': falta {', '.join(faltan)} (se omite)."); continue
        sub = df[[cD, cC, cT]].copy(); sub.columns = ["DNI", "CLIENTE", "CELULAR"]
        sub["TARJETA_WF"] = df[cK].astype(str) if cK else ""
        sub["FECHA_FORM"] = df[cF] if cF else pd.NaT
        sub["__hoja"] = sh
        frames.append(sub)
    if not frames: return None, avisos
    return pd.concat(frames, ignore_index=True), avisos

def dni_activadas(libro, hojas):
    s = set()
    for sh in hojas:
        df = libro[sh]; cD = _match_col(df.columns, ALIAS["DNI"])
        if cD: s |= set(normaliza_dni(df[cD]).replace("", pd.NA).dropna())
    return s

def cruzar(form_df, set_activadas):
    """Excluye activados, valida, dedup por DNI (conserva orden). Devuelve TODA la base no activada."""
    df = form_df.copy()
    df["DNI_N"] = normaliza_dni(df["DNI"])
    total_form = df["DNI_N"].replace("", pd.NA).dropna().nunique()
    df = df[~df["DNI_N"].isin(set_activadas)]
    df = df[df["DNI_N"].ne("")]
    df["CEL_N"] = df["CELULAR"].apply(normaliza_celular)
    df = df[df["CEL_N"].ne("")]
    df["PRIMER_NOMBRE"] = df["CLIENTE"].apply(primer_nombre)
    df = df[df["PRIMER_NOMBRE"].ne("")]
    df["ES_CERO"] = df["TARJETA_WF"].astype(str).str.upper().str.contains("CERO")
    df["FECHA_FORM"] = pd.to_datetime(df["FECHA_FORM"], errors="coerce", dayfirst=True)
    df = df.drop_duplicates(subset="DNI_N", keep="first").reset_index(drop=True)
    stats = {"formalizadas": total_form, "activadas": len(set_activadas), "no_activadas": len(df)}
    return df, stats

# Buckets de antigüedad (días desde la formalización)
BUCKETS = ["0 a 3 días", "4 a 15 días", "16 a 30 días", "31 a más días", "Sin fecha"]

def bucket_antiguedad(dias):
    if pd.isna(dias): return "Sin fecha"
    d = int(dias)
    if d < 0: return "Sin fecha"
    if d <= 3: return "0 a 3 días"
    if d <= 15: return "4 a 15 días"
    if d <= 30: return "16 a 30 días"
    return "31 a más días"


# ============================================================================
# PROMOCIONES (solo datos del PPT Arrancón Agosto 2026; sin inventar cifras)
#   psi=True  -> "Pagos Sin Intereses": NO aplica a VISA CERO
# ============================================================================
def _d(dd, mm, yyyy=2026): return date(yyyy, mm, dd)

PROMOCIONES = [
    {"id":"puntos_premios","nombre":"Puntos BBVA y Grandes Premios","categoria":"Puntos / Sorteos (Altas TC)",
     "beneficio":"Hasta 18,000 Puntos BBVA + sorteo de 10 mil Puntos semanales y premios",
     "inicio":_d(1,8),"fin":_d(31,8),"no_cero":True,"publico":"Nuevas altas de TC","comercios":[],
     "condiciones":["Nuevas altas de TC entran a sorteo de 10 mil Puntos semanales",
                    "Adicional/Efectivo/Subrogado Cuotas entran a sorteo de premios",
                    "2 ganadores por altas semanales · Sorteo 4 de septiembre"],
     "restricciones":["Sujeto a altas semanales","Premios referenciales","No aplica a VISA CERO"],
     "relevancia_activacion":20,"fuerza_beneficio":25,"contexto":6,
     "speeches":["activa tu TC y participa por hasta 18,000 Puntos BBVA y premios semanales",
                 "tu nueva TC entra al sorteo de 10mil Puntos BBVA cada semana"],
     "cuando_usar":"🎯 Gancho principal de activacion durante TODO agosto. 💡 Reforzar antes de cada corte semanal del sorteo.",
     "fuente":"PPT Puntos BBVA (01-31 ago)"},

    {"id":"bono_semanal","nombre":"Bono Semanal de Puntos","categoria":"Puntos",
     "beneficio":"1,000 Puntos BBVA por semana (+1,000 al completar 4 semanas)",
     "inicio":_d(3,8),"fin":_d(31,8),"no_cero":True,"publico":"Titulares TC","comercios":[],
     "condiciones":["Cada compra desde S/500 por semana = 1k Puntos",
                    "Si cumple las 4 semanas, +1k Puntos","Consumos del 3 al 31 de agosto"],
     "restricciones":["Requiere consumo minimo semanal de S/500","No aplica a VISA CERO"],
     "relevancia_activacion":18,"fuerza_beneficio":20,"contexto":5,
     "speeches":["gana 1,000 Puntos BBVA por semana con compras desde S/500 hasta el 31/08",
                 "suma 1k Puntos BBVA cada semana y 1k mas al completar el mes"],
     "cuando_usar":"🎯 Empujar al INICIO de cada semana para que alcancen el consumo de S/500. 💡 Mayor efecto en quincena y fin de mes por liquidez.",
     "fuente":"PPT Bono Semanal (03-31 ago)"},

    {"id":"dia_nino","nombre":"Dia del Nino — Promos","categoria":"Retail / Fecha especial",
     "beneficio":"Hasta 20% dcto (Tai Loy, Mercado Libre, Coolbox, Fun Jungle Kids)",
     "inicio":_d(10,8),"fin":_d(16,8),"no_cero":False,"publico":"Titulares TC",
     "comercios":["Tai Loy 20% (desde S/80, 12-15 ago)","Mercado Libre 15% (desde S/129)",
                  "Coolbox S/80 (desde S/999)","Fun Jungle Kids S/25.9"],
     "condiciones":["Del 10 al 16 de agosto","Montos y comercios especificos por marca"],
     "restricciones":["Ventana corta (10-16 ago)","Montos minimos por comercio","Tai Loy solo 12-15 ago"],
     "relevancia_activacion":12,"fuerza_beneficio":20,"contexto":8,
     "speeches":["por el Dia del Nino aprovecha hasta 20% dcto con tu TC BBVA solo hasta el 16/08",
                 "Dia del Nino: dctos en Tai Loy, Mercado Libre y mas con tu TC hasta el 16/08"],
     "cuando_usar":"🟠 Usar YA: ventana corta (vence 16/08). Maxima urgencia estos dias.",
     "fuente":"PPT Dia del Nino (10-16 ago)"},

    {"id":"exoneracion","nombre":"Exoneracion de Membresia","categoria":"Beneficio de activación",
     "beneficio":"Exoneración de membresía por 12 meses (según campaña)",
     "inicio":None,"fin":None,"no_cero":False,"publico":"Altas Nómina/Digital/Hipotecario","comercios":[],
     "condiciones":["Nómina: exonerada mientras se mantenga la marca PH",
                    "Digital: 12 meses · activación hasta el 18 del mes siguiente",
                    "Hipotecario: hasta el término del crédito"],
     "restricciones":["Fecha máxima de activación: 18 del mes siguiente (Digital)"],
     "relevancia_activacion":20,"fuerza_beneficio":17,"contexto":5,
     "speeches":["activa tu TC y disfruta exoneracion de membresia por 12 meses",
                 "estrena tu TC con membresia exonerada; actívala antes del 18"],
     "cuando_usar":"🎯 Util a FIN de mes para cerrar activaciones antes del limite '18 del mes siguiente'.",
     "fuente":"PPT Exoneraciones de membresía"},

    {"id":"educacion","nombre":"Educacion — Matriculas","categoria":"Educación","no_cero":True,
     "beneficio":"12 Pagos Sin Intereses en universidades (+x5 Puntos en el extranjero)",
     "inicio":None,"fin":None,"publico":"Titulares TC con pagos educativos",
     "comercios":["PUCP","USIL","UPN","y otras universidades afiliadas"],
     "condiciones":["12 Pagos Sin Intereses en matrículas","x5 Puntos por pagos en universidades en el extranjero"],
     "restricciones":["Aplica en universidades afiliadas","Vigencia no especificada en el PPT",
                      "No aplica a VISA CERO (Pagos Sin Intereses)"],
     "relevancia_activacion":12,"fuerza_beneficio":18,"contexto":7,
     "speeches":["paga tu matricula universitaria en 12 Pagos Sin Intereses con tu TC BBVA",
                 "matriculas al dia: 12 Pagos Sin Intereses en tu universidad con BBVA"],
     "cuando_usar":"💡 Matriculas se concentran al INICIO de ciclo academico; cuotas suelen caer en quincena/fin de mes. Recomendar cerca de esas fechas. (No VISA CERO).",
     "fuente":"PPT Educacion (sin fecha explícita)"},

    {"id":"exterior","nombre":"x5 Puntos en el Exterior","categoria":"Puntos / Viajes","no_cero":True,
     "beneficio":"x5 Puntos BBVA en compras desde $50 en el exterior",
     "inicio":None,"fin":None,"publico":"Titulares y adicionales","comercios":[],
     "condiciones":["Cada compra desde $50 multiplica x5 los Puntos","TC adicionales también aplican",
                    "Compras online o presenciales en el extranjero"],
     "restricciones":["Aplica a compras en el exterior","Vigencia no especificada en el PPT","No aplica a VISA CERO"],
     "relevancia_activacion":12,"fuerza_beneficio":16,"contexto":4,
     "speeches":["multiplica x5 tus Puntos BBVA en tus compras desde $50 en el exterior",
                 "viaja y suma x5 Puntos BBVA con tu TC en el extranjero"],
     "cuando_usar":"💡 Relevante en temporada de viajes, feriados y fines de semana largos.",
     "fuente":"PPT x5 exterior (sin fecha explícita)"},

    {"id":"deporte","nombre":"Semana del Deporte","categoria":"Deporte","no_cero":True,
     "beneficio":"6 Pagos Sin Intereses en marcas deportivas",
     "inicio":None,"fin":None,"publico":"Titulares TC",
     "comercios":["Reebok","Puma","Nike","Under Armour","Marathon"],
     "condiciones":["6 Pagos Sin Intereses","Promociones exclusivas en marcas deportivas"],
     "restricciones":["Aplica en marcas participantes","Vigencia no especificada en el PPT",
                      "No aplica a VISA CERO (Pagos Sin Intereses)"],
     "relevancia_activacion":12,"fuerza_beneficio":15,"contexto":4,
     "speeches":["vuelve la Semana del Deporte: 6 Pagos Sin Intereses en Nike, Puma y mas",
                 "equipate con 6 Pagos Sin Intereses en las mejores marcas deportivas"],
     "cuando_usar":"🎯 'Semana del Deporte'; sin fecha exacta en el PPT. Usar como refuerzo. (No VISA CERO).",
     "fuente":"PPT Deporte (sin fecha explícita)"},

    {"id":"conciertos","nombre":"Conciertos","categoria":"Entretenimiento","no_cero":False,
     "beneficio":"15% de descuento en eventos",
     "inicio":None,"fin":None,"publico":"Titulares TC",
     "comercios":["Black Eyed Peas","Soy Luna","Circo Mistico del Condor","Iron Maiden"],
     "condiciones":["15% de descuento en eventos favoritos","Campañas para ganar entradas dobles"],
     "restricciones":["Aplica en eventos participantes","Vigencia no especificada en el PPT"],
     "relevancia_activacion":12,"fuerza_beneficio":15,"contexto":5,
     "speeches":["disfruta tus conciertos favoritos con 15% de dcto usando tu TC BBVA",
                 "15% de dcto en tus eventos favoritos con tu Tarjeta BBVA"],
     "cuando_usar":"💡 La compra de entradas suele concentrarse cerca de la fecha del evento y en quincena/fin de mes por liquidez.",
     "fuente":"PPT Conciertos (sin fecha explícita)"},
]


# ============================================================================
# ANÁLISIS: estado / días / score
# ============================================================================
def estado_promocion(promo, hoy):
    ini, fin = promo["inicio"], promo["fin"]
    if ini is None and fin is None:
        return {"estado":"Continua","dias_restantes":None,"pill":"green",
                "detalle":"Vigencia no especificada en el PPT (campaña continua)"}
    if fin and hoy > fin:
        return {"estado":"Vencida","dias_restantes":(fin-hoy).days,"pill":"red",
                "detalle":f"Estuvo vigente del {ini:%d/%m/%Y} al {fin:%d/%m/%Y}"}
    if ini and hoy < ini:
        return {"estado":"Proxima a iniciar","dias_restantes":(ini-hoy).days,"pill":"yellow",
                "detalle":f"Inicia el {ini:%d/%m/%Y}"}
    dias = (fin-hoy).days if fin else None
    if dias is not None and dias <= 5:
        return {"estado":"Proxima a vencer","dias_restantes":dias,"pill":"orange","detalle":f"Vence el {fin:%d/%m/%Y}"}
    return {"estado":"Activa","dias_restantes":dias,"pill":"green",
            "detalle":(f"Vigente hasta el {fin:%d/%m/%Y}" if fin else "Vigente")}

def score_promocion(promo, est):
    d = est["dias_restantes"]
    if est["estado"] == "Vencida": urg = 0
    elif est["estado"] == "Proxima a iniciar": urg = 8
    elif d is None: urg = 15
    elif d <= 3: urg = 40
    elif d <= 7: urg = 32
    elif d <= 15: urg = 22
    else: urg = 14
    penal = 0
    fuertes = [r for r in promo["restricciones"] if any(k in r.lower()
               for k in ["minimo","monto","corta","solo","afiliad","participant"])]
    if fuertes: penal = min(8, 3 + len(fuertes)*2)
    total = max(0, min(100, urg + promo["fuerza_beneficio"] + promo["relevancia_activacion"] + promo["contexto"] - penal))
    return total

def analizar(hoy):
    filas = []
    for p in PROMOCIONES:
        est = estado_promocion(p, hoy)
        filas.append({**p, "_est": est, "_score": score_promocion(p, est)})
    filas.sort(key=lambda x: x["_score"], reverse=True)
    return filas


# ============================================================================
# MENSAJES
# ============================================================================
def construir_mensaje(nombre, cuerpo):
    cuerpo = cuerpo.strip().rstrip(".")
    txt = f"{nombre}, {cuerpo}. {CIERRE_OBLIGATORIO}"
    return quitar_tildes(re.sub(r"\s+", " ", txt).strip())

def validar_mensaje(msg):
    n = len(msg)
    ok = n <= LIMITE_CARACTERES and msg.strip().endswith(quitar_tildes(CIERRE_OBLIGATORIO)) and "," in msg
    return {"ok": ok, "n": n, "exceso": max(0, n - LIMITE_CARACTERES)}

def espacio_disponible(nombre_mas_largo):
    fijo = nombre_mas_largo + len(", ") + len(". ") + len(quitar_tildes(CIERRE_OBLIGATORIO))
    return LIMITE_CARACTERES - fijo

def elegir_speech(promo, nombre="NOMBRE"):
    for c in promo["speeches"]:
        m = construir_mensaje(nombre, c)
        if validar_mensaje(m)["ok"]: return c, m
    return promo["speeches"][0], construir_mensaje(nombre, promo["speeches"][0])


# ============================================================================
# EXCEL (formato exacto plantilla sms: REC OTROS, cabecera plana,
#        Telefono numerico General, DNI texto @, anchos plantilla)
# ============================================================================
def generar_excel(df_out):
    out = df_out[["Telefono", "Mensaje", "DNI"]].copy()
    out["Telefono"] = pd.to_numeric(out["Telefono"], errors="coerce").astype("Int64")
    out["DNI"] = out["DNI"].astype(str)
    buff = io.BytesIO()
    with pd.ExcelWriter(buff, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name=HOJA_SALIDA, index=False)
        ws = xw.sheets[HOJA_SALIDA]
        for cell in ws["A"][1:]:  # Telefono numerico
            cell.number_format = "General"
        for cell in ws["C"][1:]:  # DNI texto
            cell.number_format = "@"
        for col, w in ANCHOS.items():
            ws.column_dimensions[col].width = w
        # Formato TABLA de Excel (encabezado azul, filtros y filas en bandas)
        ultima = ws.max_row if ws.max_row >= 2 else 2
        tabla = Table(displayName="REC_OTROS", ref=f"A1:C{ultima}")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws.add_table(tabla)
    return buff.getvalue()


# ============================================================================
# UI
# ============================================================================
st.markdown(f"""
<div class="pm-header">
  <div><div class="pm-logo">Plus<span>Metas</span></div>
  <div class="sub">Generador Inteligente de Campañas de Activación · BBVA — Telemarketing</div></div>
  <div class="pm-badge">Arrancón · Agosto 2026</div>
</div>""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    hoy = st.date_input("Fecha de trabajo (para vigencias)", value=date.today(), format="DD/MM/YYYY")
    st.caption("Prefijo telefónico de salida: **51** (ej. 51999999999)")
    st.caption("Nombre: formato BBVA (APELLIDO APELLIDO NOMBRE).")
    st.divider(); st.caption("PlusMetas · MF Asesoría y Consultoría")

# PASO 1 — carga (manual)
st.markdown('<div class="sectionbar">📁 Paso 1 · Cargar bases</div>', unsafe_allow_html=True)
c1, c2 = st.columns(2)
with c1:
    up_form = st.file_uploader("📗 FORMALIZADAS (Excel)", type=["xlsx","xls"], key="form")
    st.caption("Ej.: BBVA-TLM · se combinan las hojas OUT + Hoja2.")
with c2:
    up_act = st.file_uploader("📕 ACTIVADAS (Excel)", type=["xlsx","xls"], key="act")
    st.caption("Ej.: ACTIVACIONES · se usa la hoja DTA JULIO.")

if not (up_form and up_act):
    st.info("Sube ambos archivos para iniciar el cruce."); st.stop()

try:
    libro_form = leer_libro(up_form.getvalue()); libro_act = leer_libro(up_act.getvalue())
except Exception:
    st.error("⚠️ No se pudo leer alguno de los archivos. Verifica que sean Excel válidos (.xlsx)."); st.stop()

val_form = hojas_validas_formalizadas(libro_form)
if not val_form:
    st.error("⚠️ En FORMALIZADAS no hay hoja con DNI + CLIENTE + CELULAR.")
    st.write("Hojas detectadas:", list(libro_form.keys())); st.stop()
hojas_dni_act = hojas_con_dni(libro_act)
if not hojas_dni_act:
    st.error("⚠️ En ACTIVADAS no hay hoja con columna DNI.")
    st.write("Hojas detectadas:", list(libro_act.keys())); st.stop()

cA, cB = st.columns(2)
with cA:
    sel_form = st.multiselect("Hojas de FORMALIZADAS a combinar", val_form, default=val_form)
with cB:
    default_act = [DEFAULT_HOJA_ACTIVADAS] if DEFAULT_HOJA_ACTIVADAS in hojas_dni_act else hojas_dni_act
    sel_act = st.multiselect("Hojas de ACTIVADAS (cruce por DNI)", hojas_dni_act, default=default_act)
if not sel_form or not sel_act:
    st.warning("Selecciona al menos una hoja en cada archivo."); st.stop()

# PASO 2 — cruce
form_df, avisos = construir_formalizadas(libro_form, sel_form)
for a in avisos: st.warning("⚠️ " + a)
if form_df is None: st.stop()
set_act = dni_activadas(libro_act, sel_act)
base_all, stats = cruzar(form_df, set_act)
if base_all.empty:
    st.error("⚠️ Tras el cruce no quedaron registros válidos."); st.stop()

st.markdown('<div class="sectionbar">🔀 Paso 2 · Cruce Formalizadas − Activadas</div>', unsafe_allow_html=True)

# Segmento de tarjeta (VISA CERO no recibe Pagos Sin Intereses)
n_cero = int(base_all["ES_CERO"].sum()); n_otros = len(base_all) - n_cero
seg = st.radio("Segmento de tarjeta (columna Tarjeta_WF)",
               ["Todos", "Tarjeta Cero (VISA CERO)", "Otros (no cero)"], horizontal=True,
               help="5 promos NO se envían a VISA CERO: Puntos y Grandes Premios, Bono Semanal, Educación, x5 Exterior, Deporte.")
if seg.startswith("Tarjeta Cero"):
    base_seg = base_all[base_all["ES_CERO"]].reset_index(drop=True)
elif seg.startswith("Otros"):
    base_seg = base_all[~base_all["ES_CERO"]].reset_index(drop=True)
else:
    base_seg = base_all

# --- Segmentación por antigüedad de formalización (días desde Fecha_Formalizacion) ---
base_seg = base_seg.copy()
base_seg["DIAS_FORM"] = (pd.Timestamp(hoy).normalize() - base_seg["FECHA_FORM"].dt.normalize()).dt.days
base_seg["BUCKET"] = base_seg["DIAS_FORM"].apply(bucket_antiguedad)
conteo = base_seg["BUCKET"].value_counts().to_dict()

st.markdown("**📅 Segmentar por antigüedad de formalización (grupo a enviar)**")
etiquetas = [f"{b} ({conteo.get(b,0)})" for b in BUCKETS]
opciones_bucket = ["Todos"] + etiquetas
sel_bucket = st.radio("Grupo por días desde formalización", opciones_bucket, horizontal=True)
if sel_bucket == "Todos":
    base_bucket = base_seg
else:
    bucket_elegido = sel_bucket.rsplit(" (", 1)[0]
    base_bucket = base_seg[base_seg["BUCKET"] == bucket_elegido].reset_index(drop=True)

cM1, cM2 = st.columns([1, 3])
with cM1:
    max_reg = st.number_input("Máximo de registros a enviar", min_value=1, value=LIMITE_REGISTROS, step=100)
base999 = base_bucket.head(int(max_reg)).copy()

k1,k2,k3,k4,k5 = st.columns(5)
for col,lbl,val,cls in [
    (k1,"Formalizadas",stats["formalizadas"],""),(k2,"Activadas",stats["activadas"],""),
    (k3,"No activadas",stats["no_activadas"],"lime"),
    (k4,"Grupo elegido",len(base_bucket),""),(k5,f"A enviar (≤{int(max_reg)})",len(base999),"lime")]:
    col.markdown(f'<div class="kpi {cls}"><div class="lbl">{lbl}</div><div class="val">{val:,}</div></div>',
                 unsafe_allow_html=True)
st.caption(f"📊 Dato observado (base cargada): VISA CERO = **{n_cero:,}** · Otros = **{n_otros:,}** · "
           f"Antigüedad → " + " · ".join(f"{b}: {conteo.get(b,0)}" for b in BUCKETS if conteo.get(b,0)))

# PASO 3 — análisis de promociones
st.markdown('<div class="sectionbar">📊 Paso 3 · Análisis de Promociones</div>', unsafe_allow_html=True)
promos = analizar(hoy)
tabla = [{"Promoción":p["nombre"],
          "Inicio":p["inicio"].strftime("%d/%m") if p["inicio"] else "—",
          "Fin":p["fin"].strftime("%d/%m") if p["fin"] else "—",
          "Estado":p["_est"]["estado"],
          "Días rest.":p["_est"]["dias_restantes"] if p["_est"]["dias_restantes"] is not None else "—",
          "VISA CERO":"No" if p["no_cero"] else "Sí",
          "Score":p["_score"]} for p in promos]
st.dataframe(pd.DataFrame(tabla), use_container_width=True, hide_index=True)

# PASO 4 — Alertas de vencimiento + recomendación por promoción
st.markdown('<div class="sectionbar">🎯 Paso 4 · Alertas y recomendación por promoción</div>', unsafe_allow_html=True)

st.markdown("**⚠️ Alertas de vencimiento**")
hay_alerta = False
for p in promos:
    e = p["_est"]
    if e["estado"] == "Proxima a vencer":
        st.markdown(f'<span class="pill orange">🟠 {p["nombre"]}</span> '
                    f'<span class="small">vence en {e["dias_restantes"]} día(s) — {e["detalle"]}. '
                    f'Priorizar mientras siga vigente.</span>', unsafe_allow_html=True); hay_alerta = True
    elif e["estado"] == "Vencida":
        st.markdown(f'<span class="pill red">🔴 {p["nombre"]}</span> '
                    f'<span class="small">vencida — {e["detalle"]}. No recomendar.</span>',
                    unsafe_allow_html=True); hay_alerta = True
    elif e["estado"] == "Proxima a iniciar":
        st.markdown(f'<span class="pill yellow">🟡 {p["nombre"]}</span> '
                    f'<span class="small">inicia en {e["dias_restantes"]} día(s) — {e["detalle"]}.</span>',
                    unsafe_allow_html=True); hay_alerta = True
if not hay_alerta:
    st.caption("Sin vencimientos próximos para la fecha seleccionada.")

st.markdown("**🧭 Cuándo conviene usar cada promoción**")
st.caption("💡 Hipótesis / 🎯 Recomendación (buenas prácticas). No hay data histórica de consumo cargada; "
           "validar con histórico antes de tomarlo como dato.")
for p in promos:
    e = p["_est"]; icon = {"red":"🔴","orange":"🟠","green":"🟢","yellow":"🟡","blue":"🔵"}[e["pill"]]
    _, ejemplo = elegir_speech(p)
    cls = "promo psi" if p["no_cero"] else "promo"
    psi_tag = ' · <b style="color:#b96a00">No VISA CERO</b>' if p["no_cero"] else ""
    st.markdown(f"""
    <div class="{cls}">
      <h4>{icon} {p['nombre']} <span class="small">· {e['estado']} · score {p['_score']}/100{psi_tag}</span></h4>
      <div class="small">{p['beneficio']} — {e['detalle']}</div>
      <div style="margin:6px 0">{p['cuando_usar']}</div>
      <code>{ejemplo}</code>
    </div>""", unsafe_allow_html=True)

# PASO 5 — selección de campaña y speech
st.markdown('<div class="sectionbar">✍️ Paso 5 · Seleccionar campaña y generar speech</div>', unsafe_allow_html=True)
vigentes = [p for p in promos if p["_est"]["estado"] not in ("Vencida","Proxima a iniciar")]

es_cero_seg = seg.startswith("Tarjeta Cero")
disponibles = [p for p in vigentes if not (es_cero_seg and p["no_cero"])]
if es_cero_seg:
    st.info("Segmento **VISA CERO**: se ocultan las promos que NO pueden enviarse a VISA CERO (Puntos y Grandes Premios, Bono Semanal, Educación, x5 Exterior, Deporte).")

opciones = {f"{p['nombre']}  ·  ({p['_est']['estado']})": p for p in disponibles}
opciones["OTRO (mensaje personalizado)"] = None
sel_key = st.selectbox("Campaña", list(opciones.keys()), index=0)
promo_sel = opciones[sel_key]

if promo_sel is None:
    cuerpo_txt = st.text_area("Escribe tu mensaje (se añade primer nombre, coma, cierre y se quitan tildes):",
                              value="aprovecha los beneficios de tu tarjeta BBVA", height=80)
else:
    cuerpos = promo_sel["speeches"]
    idx = st.radio("Speech base", list(range(len(cuerpos))), format_func=lambda i: cuerpos[i], index=0)
    cuerpo_txt = cuerpos[idx]

# Construcción de la base del envío (protección PSI vs VISA CERO en modo "Todos")
run = base999.copy()
nota_excluidos = ""
if promo_sel is not None and promo_sel["no_cero"] and not es_cero_seg and seg == "Todos":
    antes = len(base_bucket)
    run = base_bucket[~base_bucket["ES_CERO"]].head(int(max_reg)).copy()
    quitados = antes - len(base_bucket[~base_bucket["ES_CERO"]])
    if quitados > 0:
        nota_excluidos = (f"Campaña no permitida para VISA CERO: se excluyeron {quitados} registro(s) del grupo. "
                          f"Base recalculada a {len(run):,} (Otros).")

nombre_max = int(run["PRIMER_NOMBRE"].str.len().max())
disp = espacio_disponible(nombre_max)
b1,b2,b3 = st.columns(3)
b1.markdown(f'<div class="kpi"><div class="lbl">Límite</div><div class="val">160</div></div>', unsafe_allow_html=True)
b2.markdown(f'<div class="kpi"><div class="lbl">Nombre más largo</div><div class="val">{nombre_max}</div></div>', unsafe_allow_html=True)
b3.markdown(f'<div class="kpi lime"><div class="lbl">Espacio p/ cuerpo</div><div class="val">{disp}</div></div>', unsafe_allow_html=True)
if nota_excluidos: st.warning("⚠️ " + nota_excluidos)

# PASO 6 — generar + validar
run["Telefono"] = run["CEL_N"]; run["DNI"] = run["DNI_N"]
run["Mensaje"] = run["PRIMER_NOMBRE"].apply(lambda n: construir_mensaje(n, cuerpo_txt))
val = run["Mensaje"].apply(validar_mensaje)
run["_ok"] = val.apply(lambda v: v["ok"]); run["_len"] = val.apply(lambda v: v["n"])
invalidos = int((~run["_ok"]).sum()); validos = int(run["_ok"].sum())

st.markdown('<div class="sectionbar">📊 Paso 6 · Resumen previo</div>', unsafe_allow_html=True)
nombre_camp = promo_sel["nombre"] if promo_sel else "OTRO"
estado_camp = promo_sel["_est"]["detalle"] if promo_sel else "Mensaje personalizado"
r1,r2,r3,r4 = st.columns(4)
r1.markdown(f'<div class="kpi"><div class="lbl">Registros</div><div class="val">{len(run):,}</div></div>', unsafe_allow_html=True)
r2.markdown(f'<div class="kpi lime"><div class="lbl">Mensajes válidos</div><div class="val">{validos:,}</div></div>', unsafe_allow_html=True)
r3.markdown(f'<div class="kpi"><div class="lbl">Inválidos (>160)</div><div class="val">{invalidos:,}</div></div>', unsafe_allow_html=True)
r4.markdown(f'<div class="kpi"><div class="lbl">Long. máx.</div><div class="val">{int(run["_len"].max())}</div></div>', unsafe_allow_html=True)
st.caption(f"Segmento: **{seg}** · Campaña: **{nombre_camp}** · {estado_camp}")
if invalidos:
    ej = run[~run["_ok"]].iloc[0]
    st.error(f"🔴 {invalidos} mensaje(s) superan 160 caracteres. Acorta el speech.")
    st.write(f"Ejemplo ({ej['_len']} chars): `{ej['Mensaje']}`")

# PASO 7 — vista previa
st.markdown('<div class="sectionbar">👁️ Paso 7 · Vista previa (primeras 20)</div>', unsafe_allow_html=True)
prev = run.head(20)[["Telefono","Mensaje","DNI","DIAS_FORM","BUCKET","_len"]].rename(
    columns={"DIAS_FORM":"Días form.","BUCKET":"Grupo","_len":"Chars"})
st.dataframe(prev, use_container_width=True, hide_index=True)

# PASO 8 — generar excel
st.markdown('<div class="sectionbar">⬇️ Paso 8 · Generar Excel</div>', unsafe_allow_html=True)
if invalidos == 0 and len(run) > 0:
    slug = re.sub(r"[^A-Z0-9]+","_", quitar_tildes(nombre_camp).upper()).strip("_")
    fname = f"{slug}_{hoy:%Y%m%d}.xlsx"
    xls_bytes = generar_excel(run)
    st.download_button(f"📥 Descargar {fname}", data=xls_bytes, file_name=fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.success(f"Listo. Hoja **{HOJA_SALIDA}** · Telefono | Mensaje | DNI · teléfono con prefijo 51 · formato tabla.")
else:
    st.button("📥 Descargar Excel", disabled=True)
    st.caption("La descarga se habilita cuando no hay mensajes inválidos.")
